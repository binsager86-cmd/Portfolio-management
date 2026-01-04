from __future__ import annotations

import argparse
import datetime as dt
import sqlite3
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


DB_DEFAULT = "portfolio.db"

# Your confirmed schema
ASSETS_TABLE = "assets"          # columns: asset_id, symbol, asset_type, exchange, currency
LEDGER_TABLE = "ledger_entries"  # columns: entry_id, entry_datetime, entry_type, asset_id, quantity, price, cash_amount, currency, notes
PRICES_TABLE = "prices"          # columns: price_id, asset_id, price_date, close_price, source
FX_TABLE = "fx_rates"            # columns: fx_id, rate_date, from_ccy, to_ccy, rate, source
SNAP_TABLE = "daily_snapshots"   # columns: snapshot_date, asset_id, ... , mkt_value_base, cost_value_base, pnl_base


BUY_TYPES = {"BUY"}
SELL_TYPES = {"SELL"}
BONUS_TYPES = {"BONUS", "BONUS_SHARES", "STOCK_DIVIDEND"}
DIV_TYPES = {"DIVIDEND", "DIV"}
CASH_IN_TYPES = {"CASH_IN", "DEPOSIT", "TOPUP", "INJECTION"}
CASH_OUT_TYPES = {"CASH_OUT", "WITHDRAWAL", "WITHDRAW", "DRAW"}


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


def parse_date(s: str) -> dt.date:
    return dt.datetime.strptime(s, "%Y-%m-%d").date()


def to_date_str(d: dt.date) -> str:
    return d.strftime("%Y-%m-%d")


def connect(db_path: str) -> sqlite3.Connection:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    return con


def col_exists(con: sqlite3.Connection, table: str, col: str) -> bool:
    rows = con.execute(f"PRAGMA table_info({table});").fetchall()
    return any(r["name"] == col for r in rows)


def get_base_currency(con: sqlite3.Connection) -> str:
    if not con.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='settings';").fetchone():
        return "KWD"
    cols = [r["name"] for r in con.execute("PRAGMA table_info(settings);").fetchall()]
    cols_l = [c.lower() for c in cols]
    if "key" in cols_l and "value" in cols_l:
        row = con.execute(
            "SELECT value FROM settings WHERE LOWER(key) IN ('base_currency','basecurrency') LIMIT 1"
        ).fetchone()
        return (row["value"] if row and row["value"] else "KWD").upper()
    if "base_currency" in cols:
        row = con.execute("SELECT base_currency FROM settings LIMIT 1").fetchone()
        return (row["base_currency"] if row and row["base_currency"] else "KWD").upper()
    return "KWD"


def fx_to_base(con: sqlite3.Connection, base: str, ccy: str, asof: dt.date) -> float:
    base = base.upper()
    ccy = ccy.upper()
    if ccy == base:
        return 1.0

    # direct
    row = con.execute(
        f"""
        SELECT rate AS r
        FROM {FX_TABLE}
        WHERE UPPER(TRIM(from_ccy)) = UPPER(TRIM(?))
          AND UPPER(TRIM(to_ccy))   = UPPER(TRIM(?))
          AND DATE(rate_date) <= DATE(?)
        ORDER BY DATE(rate_date) DESC
        LIMIT 1;
        """,
        (ccy, base, to_date_str(asof)),
    ).fetchone()
    if row and row["r"] is not None:
        return float(row["r"])

    # inverse
    row = con.execute(
        f"""
        SELECT rate AS r
        FROM {FX_TABLE}
        WHERE UPPER(TRIM(from_ccy)) = UPPER(TRIM(?))
          AND UPPER(TRIM(to_ccy))   = UPPER(TRIM(?))
          AND DATE(rate_date) <= DATE(?)
        ORDER BY DATE(rate_date) DESC
        LIMIT 1;
        """,
        (base, ccy, to_date_str(asof)),
    ).fetchone()
    if row and row["r"] is not None and float(row["r"]) != 0:
        return 1.0 / float(row["r"])

    # fallback
    return 1.0


def price_asof(con: sqlite3.Connection, asset_id: int, asof: dt.date) -> Optional[float]:
    row = con.execute(
        f"""
        SELECT close_price AS px
        FROM {PRICES_TABLE}
        WHERE asset_id = ?
          AND DATE(price_date) <= DATE(?)
        ORDER BY DATE(price_date) DESC
        LIMIT 1;
        """,
        (asset_id, to_date_str(asof)),
    ).fetchone()
    return float(row["px"]) if row and row["px"] is not None else None


@dataclass
class Asset:
    asset_id: int
    symbol: str
    currency: str
    asset_type: str
    exchange: str


def load_assets(con: sqlite3.Connection) -> Dict[int, Asset]:
    rows = con.execute(
        "SELECT asset_id, symbol, currency, asset_type, exchange FROM assets;"
    ).fetchall()
    out: Dict[int, Asset] = {}
    for r in rows:
        out[int(r["asset_id"])] = Asset(
            asset_id=int(r["asset_id"]),
            symbol=r["symbol"] or f"ASSET_{r['asset_id']}",
            currency=(r["currency"] or "KWD").upper(),
            asset_type=r["asset_type"] or "",
            exchange=r["exchange"] or "",
        )
    return out


def detect_ledger_cols(con: sqlite3.Connection) -> Dict[str, str]:
    cols = [r["name"] for r in con.execute(f"PRAGMA table_info({LEDGER_TABLE});").fetchall()]
    m = {c.lower(): c for c in cols}

    def pick(*cands: str) -> Optional[str]:
        for c in cands:
            if c.lower() in m:
                return m[c.lower()]
        return None

    return {
        "id": pick("entry_id", "id") or "entry_id",
        "dt": pick("entry_datetime", "txn_date", "trade_date", "date") or "entry_datetime",
        "type": pick("entry_type", "txn_type", "type") or "entry_type",
        "asset": pick("asset_id") or "asset_id",
        "qty": pick("quantity", "qty", "shares") or "quantity",
        "price": pick("price", "unit_price") or "price",
        "cash": pick("cash_amount") or "cash_amount",
        "ccy": pick("currency") or "currency",
        "notes": pick("notes") or "notes",
    }


def fetch_ledger(con: sqlite3.Connection, end: dt.date) -> List[sqlite3.Row]:
    lc = detect_ledger_cols(con)
    sql = f"""
    SELECT
      {lc['id']}    AS entry_id,
      {lc['dt']}    AS entry_datetime,
      UPPER(TRIM({lc['type']})) AS entry_type,
      {lc['asset']} AS asset_id,
      {lc['qty']}   AS quantity,
      {lc['price']} AS price,
      {lc['cash']}  AS cash_amount,
      {lc['ccy']}   AS currency,
      {lc['notes']} AS notes
    FROM {LEDGER_TABLE}
    WHERE DATE({lc['dt']}) <= DATE(?)
    ORDER BY DATE({lc['dt']}), entry_id;
    """
    return con.execute(sql, (to_date_str(end),)).fetchall()


@dataclass
class RealizedRow:
    entry_id: int
    date: dt.date
    asset_id: int
    symbol: str
    qty: float
    sell_price: float
    ccy: str
    fx: float
    proceeds_base: float
    cost_basis_base: float
    realized_base: float
    realized_pct: Optional[float]


@dataclass
class TxnRow:
    entry_id: int
    date: dt.date
    entry_type: str
    asset_id: Optional[int]
    symbol: str
    qty: Optional[float]
    price: Optional[float]
    ccy: str
    fx: float
    cash_amount: Optional[float]
    cash_amount_base: Optional[float]
    proceeds_base: Optional[float]
    cost_base: Optional[float]
    realized_pnl_base: Optional[float]
    realized_pnl_pct: Optional[float]
    notes: str


def compute_realized_and_txn_pnl(
    con: sqlite3.Connection,
    assets: Dict[int, Asset],
    base: str,
    end: dt.date,
    asof_for_mtm: dt.date
) -> Tuple[List[RealizedRow], List[TxnRow], float, float, float, float]:
    """
    Average-cost realized PnL on SELL transactions.
    Also returns enriched transaction rows.
    Totals: realized_pnl_base, realized_cost_basis_base, cash_in_base, cash_out_base
    """
    ledger = fetch_ledger(con, end)

    # Running position per asset in asset currency
    qty: Dict[int, float] = {}
    cost: Dict[int, float] = {}  # total cost in asset currency

    realized_rows: List[RealizedRow] = []
    txn_rows: List[TxnRow] = []

    total_realized = 0.0
    total_realized_cost_basis = 0.0
    cash_in_base = 0.0
    cash_out_base = 0.0

    for r in ledger:
        entry_id = int(r["entry_id"])
        d = parse_date(str(r["entry_datetime"])[:10])  # handles datetime text
        t = str(r["entry_type"] or "").upper().strip()
        aid = r["asset_id"]
        aid_int = int(aid) if aid is not None else None

        a = assets.get(aid_int) if aid_int is not None else None
        sym = a.symbol if a else (f"ASSET_{aid_int}" if aid_int is not None else "")
        ccy = (str(r["currency"]) if r["currency"] else (a.currency if a else base)).upper()

        fx = fx_to_base(con, base, ccy, d)

        q = float(r["quantity"]) if r["quantity"] is not None else None
        px = float(r["price"]) if r["price"] is not None else None
        cash_amt = float(r["cash_amount"]) if r["cash_amount"] is not None else None
        cash_amt_base = (cash_amt * fx) if cash_amt is not None else None

        proceeds_base = None
        cost_base = None
        realized_base = None
        realized_pct = None

        if t in CASH_IN_TYPES and cash_amt is not None:
            cash_in_base += abs(cash_amt_base or 0.0)

        if t in CASH_OUT_TYPES and cash_amt is not None:
            cash_out_base += abs(cash_amt_base or 0.0)

        if aid_int is not None:
            qty.setdefault(aid_int, 0.0)
            cost.setdefault(aid_int, 0.0)

        # BUY
        if t in BUY_TYPES and aid_int is not None and q is not None and px is not None:
            qty[aid_int] += q
            cost[aid_int] += q * px
            cost_base = q * px * fx  # cash outflow magnitude

        # BONUS: qty increases, cost unchanged
        elif t in BONUS_TYPES and aid_int is not None and q is not None:
            qty[aid_int] += q

        # SELL: realized = (sell_px - avg_cost)*q
        elif t in SELL_TYPES and aid_int is not None and q is not None and px is not None:
            if qty.get(aid_int, 0.0) <= 0:
                # can't compute; still record txn row with blanks
                pass
            else:
                avg_cost = (cost[aid_int] / qty[aid_int]) if qty[aid_int] != 0 else 0.0
                sold_cost = avg_cost * q
                sold_proceeds = px * q
                realized = sold_proceeds - sold_cost

                # update book
                qty[aid_int] -= q
                cost[aid_int] -= sold_cost

                proceeds_base = sold_proceeds * fx
                cost_basis_base = sold_cost * fx
                realized_base = realized * fx

                total_realized += realized_base
                total_realized_cost_basis += cost_basis_base

                realized_pct = (realized_base / cost_basis_base) if cost_basis_base != 0 else None

                realized_rows.append(
                    RealizedRow(
                        entry_id=entry_id,
                        date=d,
                        asset_id=aid_int,
                        symbol=sym,
                        qty=q,
                        sell_price=px,
                        ccy=ccy,
                        fx=fx,
                        proceeds_base=proceeds_base,
                        cost_basis_base=cost_basis_base,
                        realized_base=realized_base,
                        realized_pct=realized_pct,
                    )
                )

        # DIVIDEND (optional): treat as cash inflow to investor (not realized gain)
        # If you want dividends included in "realized", we can add it, but by finance convention it’s separate.

        txn_rows.append(
            TxnRow(
                entry_id=entry_id,
                date=d,
                entry_type=t,
                asset_id=aid_int,
                symbol=sym,
                qty=q,
                price=px,
                ccy=ccy,
                fx=fx,
                cash_amount=cash_amt,
                cash_amount_base=cash_amt_base,
                proceeds_base=proceeds_base,
                cost_base=cost_base,
                realized_pnl_base=realized_base,
                realized_pnl_pct=realized_pct,
                notes=str(r["notes"] or ""),
            )
        )

    return realized_rows, txn_rows, total_realized, total_realized_cost_basis, cash_in_base, cash_out_base


def fetch_holdings_from_snapshots(con: sqlite3.Connection, asof: dt.date) -> List[sqlite3.Row]:
    return con.execute(
        f"""
        SELECT
          ds.snapshot_date,
          ds.asset_id,
          a.symbol,
          a.asset_type,
          a.exchange,
          ds.quantity,
          ds.avg_cost,
          ds.mkt_price,
          ds.currency,
          ds.fx_to_base,
          ds.cost_value_base,
          ds.mkt_value_base,
          ds.pnl_base
        FROM {SNAP_TABLE} ds
        LEFT JOIN {ASSETS_TABLE} a ON a.asset_id = ds.asset_id
        WHERE ds.snapshot_date = ?
        ORDER BY ds.mkt_value_base DESC;
        """,
        (to_date_str(asof),),
    ).fetchall()


def fetch_portfolio_value_series(con: sqlite3.Connection, start: dt.date, end: dt.date) -> List[sqlite3.Row]:
    return con.execute(
        f"""
        SELECT snapshot_date, SUM(mkt_value_base) AS portfolio_value_base
        FROM {SNAP_TABLE}
        WHERE snapshot_date BETWEEN ? AND ?
        GROUP BY snapshot_date
        ORDER BY snapshot_date;
        """,
        (to_date_str(start), to_date_str(end)),
    ).fetchall()


def autosize_and_format_table(ws, header_row: int = 1, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    # header style
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # autosize
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    ws.auto_filter.ref = ws.dimensions


def write_sheet(ws, title: str, headers: List[str], rows: List[List[object]]) -> None:
    ws.title = title
    ws.append(headers)
    for r in rows:
        ws.append(r)
    autosize_and_format_table(ws)


def export_excel(db_path: str, asof: dt.date, start: Optional[dt.date], end: Optional[dt.date], out_path: str) -> None:
    con = connect(db_path)
    base = get_base_currency(con)
    assets = load_assets(con)

    # Require snapshots for holdings/series
    if not con.execute(f"SELECT 1 FROM sqlite_master WHERE type='table' AND name='{SNAP_TABLE}';").fetchone():
        raise RuntimeError("daily_snapshots table is missing. Build snapshots first with daily_snapshot.py")

    holdings = fetch_holdings_from_snapshots(con, asof)
    if not holdings:
        raise RuntimeError(f"No snapshot rows found for {to_date_str(asof)}. Build snapshots for that date first.")

    # Date range for series (defaults)
    if start is None or end is None:
        # fallback to available snapshot min/max
        r = con.execute(f"SELECT MIN(snapshot_date) AS mn, MAX(snapshot_date) AS mx FROM {SNAP_TABLE};").fetchone()
        start = parse_date(r["mn"]) if r and r["mn"] else asof
        end = parse_date(r["mx"]) if r and r["mx"] else asof

    series = fetch_portfolio_value_series(con, start, end)

    # Realized and per-transaction PnL
    realized_rows, txn_rows, realized_total, realized_cost_basis_total, cash_in_base, cash_out_base = compute_realized_and_txn_pnl(
        con, assets, base, end, asof
    )

    # Unrealized totals from holdings (base)
    unrealized_total = sum(float(r["pnl_base"]) for r in holdings)
    cost_total = sum(float(r["cost_value_base"]) for r in holdings)
    mkt_total = sum(float(r["mkt_value_base"]) for r in holdings)
    unrealized_pct = (unrealized_total / cost_total) if cost_total != 0 else None
    realized_pct = (realized_total / realized_cost_basis_total) if realized_cost_basis_total != 0 else None

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Summary sheet ----
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Portfolio Summary"
    ws["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ["As Of", to_date_str(asof)],
        ["Base Currency", base],
        ["Holdings Cost (Base)", cost_total],
        ["Holdings Market Value (Base)", mkt_total],
        ["Unrealized PnL (Base)", unrealized_total],
        ["Unrealized PnL %", unrealized_pct],
        ["Realized PnL (Base) [SELLs]", realized_total],
        ["Realized PnL % [SELLs]", realized_pct],
        ["Cash In (Base)", cash_in_base],
        ["Cash Out (Base)", cash_out_base],
    ]
    ws.append(["Metric", "Value"])
    for r in summary_rows:
        ws.append(r)

    # formatting summary
    autosize_and_format_table(ws, header_row=2, freeze="A3")
    # number formats
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell.value, float):
            cell.number_format = "#,##0.000"
    # percent rows
    for i in range(3, ws.max_row + 1):
        if ws[f"A{i}"].value in ("Unrealized PnL %", "Realized PnL % [SELLs]") and ws[f"B{i}"].value is not None:
            ws[f"B{i}"].number_format = "0.00%"

    # ---- Holdings sheet ----
    ws = wb.create_sheet("Holdings")
    headers = [
        "snapshot_date", "asset_id", "symbol", "asset_type", "exchange",
        "quantity", "avg_cost", "mkt_price", "currency", "fx_to_base",
        "cost_value_base", "mkt_value_base", "unrealized_pnl_base", "unrealized_pnl_%"
    ]
    rows = []
    for r in holdings:
        pnl = float(r["pnl_base"])
        costb = float(r["cost_value_base"])
        pct = (pnl / costb) if costb != 0 else None
        rows.append([
            r["snapshot_date"], r["asset_id"], r["symbol"], r["asset_type"], r["exchange"],
            float(r["quantity"]), float(r["avg_cost"]), float(r["mkt_price"]), r["currency"], float(r["fx_to_base"]),
            float(r["cost_value_base"]), float(r["mkt_value_base"]), pnl, pct
        ])
    write_sheet(ws, "Holdings", headers, rows)
    # formats
    for i in range(2, ws.max_row + 1):
        ws[f"F{i}"].number_format = "#,##0.########"
        ws[f"G{i}"].number_format = "#,##0.000000"
        ws[f"H{i}"].number_format = "#,##0.000000"
        ws[f"J{i}"].number_format = "0.000000"
        for col in ("K", "L", "M"):
            ws[f"{col}{i}"].number_format = "#,##0.000"
        if ws[f"N{i}"].value is not None:
            ws[f"N{i}"].number_format = "0.00%"

    # ---- Transactions sheet ----
    ws = wb.create_sheet("Transactions")
    headers = [
        "entry_id", "date", "type", "asset_id", "symbol",
        "quantity", "price", "currency", "fx_to_base",
        "cash_amount", "cash_amount_base",
        "buy_cost_base", "sell_proceeds_base",
        "realized_pnl_base", "realized_pnl_%", "notes"
    ]
    rows = []
    for t in txn_rows:
        rows.append([
            t.entry_id, to_date_str(t.date), t.entry_type, t.asset_id, t.symbol,
            t.qty, t.price, t.ccy, t.fx,
            t.cash_amount, t.cash_amount_base,
            t.cost_base, t.proceeds_base,
            t.realized_pnl_base, t.realized_pnl_pct, t.notes
        ])
    write_sheet(ws, "Transactions", headers, rows)
    for i in range(2, ws.max_row + 1):
        ws[f"B{i}"].number_format = "yyyy-mm-dd"
        for col in ("F",):
            ws[f"{col}{i}"].number_format = "#,##0.########"
        for col in ("G", "I"):
            ws[f"{col}{i}"].number_format = "#,##0.000000"
        for col in ("J", "K", "L", "M", "N"):
            ws[f"{col}{i}"].number_format = "#,##0.000"
        if ws[f"O{i}"].value is not None:
            ws[f"O{i}"].number_format = "0.00%"

    # ---- Realized sheet ----
    ws = wb.create_sheet("Realized")
    headers = [
        "entry_id", "date", "asset_id", "symbol", "qty",
        "sell_price", "currency", "fx_to_base",
        "proceeds_base", "cost_basis_base", "realized_pnl_base", "realized_pnl_%"
    ]
    rows = []
    for rr in realized_rows:
        rows.append([
            rr.entry_id, to_date_str(rr.date), rr.asset_id, rr.symbol, rr.qty,
            rr.sell_price, rr.ccy, rr.fx,
            rr.proceeds_base, rr.cost_basis_base, rr.realized_base, rr.realized_pct
        ])
    write_sheet(ws, "Realized", headers, rows)
    for i in range(2, ws.max_row + 1):
        ws[f"B{i}"].number_format = "yyyy-mm-dd"
        ws[f"E{i}"].number_format = "#,##0.########"
        ws[f"F{i}"].number_format = "#,##0.000000"
        ws[f"H{i}"].number_format = "#,##0.000000"
        for col in ("I", "J", "K"):
            ws[f"{col}{i}"].number_format = "#,##0.000"
        if ws[f"L{i}"].value is not None:
            ws[f"L{i}"].number_format = "0.00%"

    # ---- Portfolio Value Series sheet ----
    ws = wb.create_sheet("PortfolioValue")
    headers = ["snapshot_date", "portfolio_value_base"]
    rows = [[r["snapshot_date"], float(r["portfolio_value_base"])] for r in series]
    write_sheet(ws, "PortfolioValue", headers, rows)
    for i in range(2, ws.max_row + 1):
        ws[f"A{i}"].number_format = "yyyy-mm-dd"
        ws[f"B{i}"].number_format = "#,##0.000"

    wb.save(out_path)
    con.close()


def main() -> None:
    ap = argparse.ArgumentParser(description="Export portfolio report to Excel (realized/unrealized/transactions).")
    ap.add_argument("--db", default=DB_DEFAULT, help="Path to SQLite DB (default: portfolio.db)")
    ap.add_argument("--asof", required=True, help="As-of date YYYY-MM-DD (must exist in daily_snapshots)")
    ap.add_argument("--start", help="Start date YYYY-MM-DD (for PortfolioValue series)")
    ap.add_argument("--end", help="End date YYYY-MM-DD (for PortfolioValue series and realized calc)")
    ap.add_argument("--out", help="Output xlsx path")
    args = ap.parse_args()

    asof = parse_date(args.asof)
    start = parse_date(args.start) if args.start else None
    end = parse_date(args.end) if args.end else None

    out = args.out or f"portfolio_export_{args.asof}.xlsx"
    export_excel(args.db, asof=asof, start=start, end=end or asof, out_path=out)
    print(f"✅ Excel exported: {out}")


if __name__ == "__main__":
    main()
